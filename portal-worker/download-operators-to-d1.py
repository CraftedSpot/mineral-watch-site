#!/usr/bin/env python3
"""
Download and parse OCC operator list to D1 database
"""
import os
import urllib.request
from openpyxl import load_workbook
from datetime import datetime

print("Oklahoma Corporation Commission Operator List Importer")
print(f"Time: {datetime.now()}")
print("=" * 60)

# Download the file
operator_list_url = "https://oklahoma.gov/content/dam/ok/en/occ/documents/og/ogdatafiles/operator-list.xlsx"
local_file = "occ-operator-list.xlsx"

print(f"\nDownloading operator list from OCC...")
print(f"URL: {operator_list_url}")

try:
    response = requests.get(operator_list_url, timeout=60)
    response.raise_for_status()
    
    with open(local_file, 'wb') as f:
        f.write(response.content)
    
    file_size_mb = len(response.content) / 1024 / 1024
    print(f"✓ Downloaded {file_size_mb:.1f} MB")
except Exception as e:
    print(f"✗ Error downloading file: {e}")
    exit(1)

# Create operators table SQL
create_table_sql = """
-- Create operators table in D1
CREATE TABLE IF NOT EXISTS operators (
    operator_number TEXT PRIMARY KEY,
    operator_name TEXT NOT NULL,
    status TEXT,
    phone TEXT,
    address TEXT,
    city TEXT,
    state TEXT,
    zip TEXT,
    contact_name TEXT,
    operator_name_normalized TEXT,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
);

-- Create index on normalized name for fast lookups
CREATE INDEX IF NOT EXISTS idx_operators_normalized ON operators(operator_name_normalized);
CREATE INDEX IF NOT EXISTS idx_operators_name ON operators(operator_name);
CREATE INDEX IF NOT EXISTS idx_operators_status ON operators(status);
"""

with open('create-operators-table.sql', 'w') as f:
    f.write(create_table_sql)

print(f"\n✓ Created table schema: create-operators-table.sql")

# Parse Excel file
print(f"\nParsing {local_file}...")

wb = load_workbook(local_file, read_only=True)
ws = wb.active

# Statistics
stats = {
    'total_rows': 0,
    'valid_operators': 0,
    'open_operators': 0,
    'closed_operators': 0,
    'skipped': 0
}

# Column mapping
columns = {}
header_row = next(ws.iter_rows(values_only=True))

for idx, header in enumerate(header_row):
    if header:
        header_clean = str(header).strip().upper()
        if 'OPERATOR NO' in header_clean or 'OPERATOR #' in header_clean:
            columns['operator_no'] = idx
        elif header_clean == 'OPERATOR':
            columns['operator_name'] = idx
        elif 'STATUS' in header_clean:
            columns['status'] = idx
        elif 'PHONE' in header_clean:
            columns['phone'] = idx
        elif 'ADDRESS' in header_clean and 'STREET' in header_clean:
            columns['address'] = idx
        elif 'CITY' in header_clean:
            columns['city'] = idx
        elif 'STATE' in header_clean:
            columns['state'] = idx
        elif 'ZIP' in header_clean:
            columns['zip'] = idx
        elif 'CONTACT' in header_clean:
            columns['contact'] = idx

print(f"\nFound columns:")
for col, idx in columns.items():
    print(f"  {col}: column {idx} ({header_row[idx]})")

# Generate SQL inserts
batch_num = 0
batch_size = 1000
current_batch = []

def escape_sql(value):
    """Escape single quotes for SQL"""
    if value is None:
        return 'NULL'
    return "'" + str(value).replace("'", "''") + "'"

def normalize_operator_name(name):
    """Normalize operator name for searching"""
    if not name:
        return ''
    return name.strip().upper().replace('.', '').replace(',', '')

def write_batch(batch_data, batch_number):
    """Write batch to SQL file"""
    filename = f'operators-batch-{str(batch_number).zfill(3)}.sql'
    
    with open(filename, 'w', encoding='utf-8') as f:
        if batch_number == 1:
            # First batch: clear existing data
            f.write("-- Clear existing operator data\n")
            f.write("DELETE FROM operators;\n\n")
        
        f.write("-- Insert operator batch\n")
        for insert in batch_data:
            f.write(insert + ";\n")
    
    print(f"  ✓ Wrote batch {batch_number}: {filename} ({len(batch_data)} operators)")

# Process data rows
print(f"\nProcessing operator data...")

for row in ws.iter_rows(min_row=2, values_only=True):
    stats['total_rows'] += 1
    
    if stats['total_rows'] % 1000 == 0:
        print(f"  Processed {stats['total_rows']:,} rows...")
    
    # Extract operator data
    operator_no = row[columns['operator_no']] if 'operator_no' in columns else None
    operator_name = row[columns['operator_name']] if 'operator_name' in columns else None
    
    if not operator_no or not operator_name:
        stats['skipped'] += 1
        continue
    
    # Clean data
    operator_no = str(operator_no).strip()
    operator_name = str(operator_name).strip()
    
    if not operator_no or not operator_name:
        stats['skipped'] += 1
        continue
    
    stats['valid_operators'] += 1
    
    # Get other fields
    status = str(row[columns['status']]).strip() if 'status' in columns and row[columns['status']] else 'UNKNOWN'
    phone = str(row[columns['phone']]).strip() if 'phone' in columns and row[columns['phone']] else None
    address = str(row[columns['address']]).strip() if 'address' in columns and row[columns['address']] else None
    city = str(row[columns['city']]).strip() if 'city' in columns and row[columns['city']] else None
    state = str(row[columns['state']]).strip() if 'state' in columns and row[columns['state']] else None
    zip_code = str(row[columns['zip']]).strip() if 'zip' in columns and row[columns['zip']] else None
    contact = str(row[columns['contact']]).strip() if 'contact' in columns and row[columns['contact']] else None
    
    # Clean phone number
    if phone and phone.upper() in ['NONE', 'NULL', 'N/A']:
        phone = None
    
    # Track status
    if status.upper() == 'OPEN':
        stats['open_operators'] += 1
    elif status.upper() == 'CLOSED':
        stats['closed_operators'] += 1
    
    # Normalize operator name for searching
    operator_name_normalized = normalize_operator_name(operator_name)
    
    # Build INSERT statement
    insert_sql = f"""INSERT INTO operators (
        operator_number,
        operator_name,
        status,
        phone,
        address,
        city,
        state,
        zip,
        contact_name,
        operator_name_normalized
    ) VALUES (
        {escape_sql(operator_no)},
        {escape_sql(operator_name)},
        {escape_sql(status)},
        {escape_sql(phone)},
        {escape_sql(address)},
        {escape_sql(city)},
        {escape_sql(state)},
        {escape_sql(zip_code)},
        {escape_sql(contact)},
        {escape_sql(operator_name_normalized)}
    )"""
    
    current_batch.append(insert_sql)
    
    # Write batch if full
    if len(current_batch) >= batch_size:
        batch_num += 1
        write_batch(current_batch, batch_num)
        current_batch = []

# Write final batch
if current_batch:
    batch_num += 1
    write_batch(current_batch, batch_num)

# Close workbook
wb.close()

# Print summary
print(f"\n{'=' * 60}")
print(f"Import Summary:")
print(f"  Total rows processed: {stats['total_rows']:,}")
print(f"  Valid operators: {stats['valid_operators']:,}")
print(f"    - Open: {stats['open_operators']:,}")
print(f"    - Closed: {stats['closed_operators']:,}")
print(f"  Skipped rows: {stats['skipped']:,}")
print(f"  SQL files generated: {batch_num}")
print(f"\nNext steps:")
print(f"  1. Run: wrangler d1 execute oklahoma-wells --remote --file=create-operators-table.sql")
print(f"  2. Import batches: ./import-operators.sh")

# Create import script
import_script = """#!/bin/bash
# Import operator data to D1

echo "Creating operators table..."
wrangler d1 execute oklahoma-wells --remote --file=create-operators-table.sql

echo "Importing operator data..."
for file in operators-batch-*.sql; do
    if [ -f "$file" ]; then
        echo "  Importing $file..."
        wrangler d1 execute oklahoma-wells --remote --file="$file" -y
    fi
done

echo "Done! Verifying import..."
wrangler d1 execute oklahoma-wells --remote --command="SELECT COUNT(*) as total, COUNT(CASE WHEN status = 'OPEN' THEN 1 END) as open FROM operators"
"""

with open('import-operators.sh', 'w') as f:
    f.write(import_script)

os.chmod('import-operators.sh', 0o755)

print(f"\n✓ Created import script: import-operators.sh")
print(f"\nDone! Total time: {datetime.now()}")