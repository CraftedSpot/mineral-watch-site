#!/usr/bin/env python3
"""
Check column headers in operator list
"""
from openpyxl import load_workbook

print("Checking operator list columns...")

wb = load_workbook('occ-operator-list.xlsx', read_only=True)
ws = wb.active

# Get header row
header_row = next(ws.iter_rows(values_only=True))

print(f"\nFound {len(header_row)} columns:")
for idx, header in enumerate(header_row):
    if header:
        print(f"  Column {idx}: {header}")

# Check first few data rows
print(f"\nFirst 3 data rows:")
row_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    row_count += 1
    if row_count > 3:
        break
    print(f"\nRow {row_count}:")
    for idx, value in enumerate(row[:15]):  # First 15 columns
        if value:
            print(f"  Col {idx}: {value}")

wb.close()