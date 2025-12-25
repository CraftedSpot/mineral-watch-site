#!/usr/bin/env python3
"""
Convert legacy completions file to SQL updates using COALESCE
to only fill in NULL values, preserving existing data
"""
from openpyxl import load_workbook

wb = load_workbook('completions-wells-legacy.xlsx', read_only=True)
ws = wb.active

batch_size = 1000
batch_num = 0
statements = []
seen_apis = set()  # Only first record per API

print("Processing legacy completions file...")
row_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    row_count += 1
    if row_count % 10000 == 0:
        print(f"Processed {row_count} rows...")
    
    api = row[0]
    if not api or api in seen_apis:
        continue
    seen_apis.add(api)
    
    mtd = row[49]  # MEAS_TOTAL_DEPTH
    tvd = row[50]  # TRUE_VERT_DEPTH
    completion = row[42]  # COMPLETION_DATE
    spud = row[43]  # SPUD_DATE
    
    # Build UPDATE for NULL fields only (don't overwrite base file data)
    updates = []
    if mtd and mtd > 0:
        updates.append(f"measured_total_depth = COALESCE(measured_total_depth, {int(mtd)})")
    if tvd and tvd > 0:
        updates.append(f"true_vertical_depth = COALESCE(true_vertical_depth, {int(tvd)})")
    if completion:
        updates.append(f"completion_date = COALESCE(completion_date, '{completion}')")
    if spud:
        updates.append(f"spud_date = COALESCE(spud_date, '{spud}')")
    
    if updates:
        sql = f"UPDATE wells SET {', '.join(updates)} WHERE api_number = '{api}';"
        statements.append(sql)
    
    if len(statements) >= batch_size:
        with open(f'legacy-batch-{batch_num:04d}.sql', 'w') as f:
            f.write('\n'.join(statements))
        print(f'Wrote legacy-batch-{batch_num:04d}.sql ({len(statements)} statements)')
        statements = []
        batch_num += 1

# Write remaining
if statements:
    with open(f'legacy-batch-{batch_num:04d}.sql', 'w') as f:
        f.write('\n'.join(statements))
    print(f'Wrote legacy-batch-{batch_num:04d}.sql ({len(statements)} statements)')
    batch_num += 1

print(f'\nTotal: {batch_num} batch files, {len(seen_apis)} unique APIs')
print(f'Total rows processed: {row_count}')