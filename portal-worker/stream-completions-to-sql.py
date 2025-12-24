#!/usr/bin/env python3
"""
Stream parse large XLSX file and generate SQL updates for well completions
"""
import os
from openpyxl import load_workbook
from datetime import datetime

print("Starting completions XLSX streaming parser...")
print(f"Time: {datetime.now()}")

# File paths
xlsx_file = 'completions-wells-formations-base.xlsx'
if not os.path.exists(xlsx_file):
    print(f"Error: {xlsx_file} not found!")
    exit(1)

print(f"\nOpening {xlsx_file} in read-only mode...")

# Statistics
stats = {
    'total_rows': 0,
    'valid_apis': 0,
    'rows_with_data': 0,
    'skipped': 0,
    'updates_generated': 0
}

# Column indices (will be determined from header row)
column_indices = {}

# SQL updates buffer
updates_buffer = []
batch_num = 1
BATCH_SIZE = 1000

def escape_sql_string(value):
    """Escape single quotes for SQL"""
    if value is None:
        return None
    return str(value).replace("'", "''")

def write_batch():
    """Write current batch to SQL file"""
    global batch_num, updates_buffer
    
    if not updates_buffer:
        return
        
    filename = f'completions-batch-{str(batch_num).zfill(4)}.sql'
    with open(filename, 'w') as f:
        f.write('\n'.join(updates_buffer))
    
    print(f"  Wrote {len(updates_buffer)} statements to {filename}")
    batch_num += 1
    updates_buffer.clear()

try:
    # Open workbook in read-only mode for streaming
    wb = load_workbook(filename=xlsx_file, read_only=True, data_only=True)
    ws = wb.active
    print(f"Opened workbook with sheet: {ws.title}")
    
    # Process rows
    for row_num, row in enumerate(ws.iter_rows(values_only=True)):
        # First row is header
        if row_num == 0:
            print("\nProcessing header row...")
            for idx, cell_value in enumerate(row):
                if cell_value:
                    col_name = str(cell_value)
                    # Map the columns we need
                    if 'API_Number' in col_name or 'API Number' in col_name:
                        column_indices['api'] = idx
                    elif 'Bottom_Hole_Lat_Y' in col_name:
                        column_indices['bh_lat'] = idx
                    elif 'Bottom_Hole_Long_X' in col_name:
                        column_indices['bh_long'] = idx
                    elif 'Formation_Name' in col_name:
                        column_indices['formation_name'] = idx
                    elif 'Formation_Depth' in col_name:
                        column_indices['formation_depth'] = idx
                    elif 'True_Vertical_Depth' in col_name:
                        column_indices['tvd'] = idx
                    elif 'Measured_Total_Depth' in col_name:
                        column_indices['total_depth'] = idx
                    elif col_name == 'Length':  # Lateral length
                        column_indices['lateral_length'] = idx
                    elif 'Well_Completion' in col_name and 'Date' not in col_name:
                        column_indices['completion_date'] = idx
                    elif 'Oil_BBL_Per_Day' in col_name:
                        column_indices['ip_oil'] = idx
                    elif 'Gas_MCF_Per_Day' in col_name:
                        column_indices['ip_gas'] = idx
                    elif 'Water_BBL_Per_Day' in col_name:
                        column_indices['ip_water'] = idx
            
            print("Column mappings found:")
            for field, idx in column_indices.items():
                print(f"  {field}: column {idx}")
            continue
        
        # Process data rows
        stats['total_rows'] += 1
        
        # Progress indicator
        if stats['total_rows'] % 10000 == 0:
            print(f"  Processed {stats['total_rows']:,} rows...")
        
        # Get API number
        if 'api' not in column_indices or len(row) <= column_indices['api']:
            stats['skipped'] += 1
            continue
            
        api = row[column_indices['api']]
        if not api:
            stats['skipped'] += 1
            continue
        
        # Clean API number (remove dashes, spaces)
        api = str(api).replace('-', '').replace(' ', '').strip()
        if not api:
            stats['skipped'] += 1
            continue
            
        stats['valid_apis'] += 1
        
        # Build UPDATE statement
        update_fields = []
        
        # Bottom hole location
        if 'bh_lat' in column_indices and len(row) > column_indices['bh_lat']:
            value = row[column_indices['bh_lat']]
            if value is not None:
                try:
                    lat = float(value)
                    if lat != 0:
                        update_fields.append(f"bh_latitude = {lat}")
                except:
                    pass
        
        if 'bh_long' in column_indices and len(row) > column_indices['bh_long']:
            value = row[column_indices['bh_long']]
            if value is not None:
                try:
                    lng = float(value)
                    if lng != 0:
                        update_fields.append(f"bh_longitude = {lng}")
                except:
                    pass
        
        # Formation name
        if 'formation_name' in column_indices and len(row) > column_indices['formation_name']:
            value = row[column_indices['formation_name']]
            if value:
                escaped = escape_sql_string(value)
                if escaped:
                    update_fields.append(f"formation_name = '{escaped}'")
        
        # Depths
        for field, col_key, db_field in [
            ('formation_depth', 'formation_depth', 'formation_depth'),
            ('tvd', 'tvd', 'true_vertical_depth'),
            ('total_depth', 'total_depth', 'measured_total_depth'),
            ('lateral_length', 'lateral_length', 'lateral_length')
        ]:
            if col_key in column_indices and len(row) > column_indices[col_key]:
                value = row[column_indices[col_key]]
                if value is not None:
                    try:
                        depth = int(float(value))
                        update_fields.append(f"{db_field} = {depth}")
                    except:
                        pass
        
        # Completion date
        if 'completion_date' in column_indices and len(row) > column_indices['completion_date']:
            value = row[column_indices['completion_date']]
            if value:
                # Handle datetime objects or strings
                try:
                    if hasattr(value, 'strftime'):
                        date_str = value.strftime('%Y-%m-%d')
                    else:
                        date_str = str(value)[:10]  # Take first 10 chars for YYYY-MM-DD
                    if date_str and date_str != 'None':
                        update_fields.append(f"completion_date = '{date_str}'")
                except:
                    pass
        
        # Initial production
        for field, col_key, db_field in [
            ('ip_oil', 'ip_oil', 'ip_oil_bbl'),
            ('ip_gas', 'ip_gas', 'ip_gas_mcf'),
            ('ip_water', 'ip_water', 'ip_water_bbl')
        ]:
            if col_key in column_indices and len(row) > column_indices[col_key]:
                value = row[column_indices[col_key]]
                if value is not None:
                    try:
                        production = float(value)
                        update_fields.append(f"{db_field} = {production}")
                    except:
                        pass
        
        # Generate UPDATE statement if we have data
        if update_fields:
            sql = f"UPDATE wells SET {', '.join(update_fields)} WHERE api_number = '{api}';"
            updates_buffer.append(sql)
            stats['rows_with_data'] += 1
            stats['updates_generated'] += 1
            
            # Write batch if buffer is full
            if len(updates_buffer) >= BATCH_SIZE:
                write_batch()
    
    # Write final batch
    write_batch()
    
    # Close workbook
    wb.close()
    
    print("\n" + "="*60)
    print("Processing complete!")
    print(f"Total rows processed: {stats['total_rows']:,}")
    print(f"Valid API numbers: {stats['valid_apis']:,}")
    print(f"Rows with completion data: {stats['rows_with_data']:,}")
    print(f"Rows skipped: {stats['skipped']:,}")
    print(f"UPDATE statements generated: {stats['updates_generated']:,}")
    print(f"Batch files created: {batch_num - 1}")
    print("="*60)
    
    # Create execution script
    if batch_num > 1:
        script_content = f"""#!/bin/bash
# Execute well completions update
echo "Starting completions data import..."
echo "Total batches: {batch_num - 1}"
echo ""

start_time=$(date +%s)
processed=0

for file in completions-batch-*.sql; do
    if [ -f "$file" ]; then
        processed=$((processed + 1))
        echo "[$(date '+%Y-%m-%d %H:%M:%S')] Processing batch $processed: $file"
        
        # Execute with wrangler
        if wrangler d1 execute oklahoma-wells --remote --file="$file"; then
            echo "  ✓ Success"
        else
            echo "  ✗ Failed - check logs"
        fi
        
        # Brief pause between batches to avoid rate limiting
        if [ $processed -lt {batch_num - 1} ]; then
            sleep 2
        fi
    fi
done

end_time=$(date +%s)
duration=$((end_time - start_time))

echo ""
echo "Completions import complete!"
echo "Total time: $duration seconds"
echo "Batches processed: $processed"
"""
        
        with open('execute-completions-import.sh', 'w') as f:
            f.write(script_content)
        
        os.chmod('execute-completions-import.sh', 0o755)
        print(f"\nCreated execution script: ./execute-completions-import.sh")
        print("\nRun the script to import completion data to D1 database")
    
except Exception as e:
    print(f"\nError: {e}")
    import traceback
    traceback.print_exc()

print(f"\nFinished at: {datetime.now()}")