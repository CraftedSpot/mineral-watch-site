#!/usr/bin/env python3
"""
Import formation data from OCC ITD (Intent to Drill) file into D1 wells table.

Only updates wells where formation_name IS NULL (COALESCE pattern).
Uses formation_normalization subquery to set formation_canonical and formation_group.

Usage:
  python3 import-itd-formations.py [--itd-file PATH]

Output:
  - itd-formation-batch-NNNN.sql files (1000 statements each)
  - execute-itd-import.sh script
"""
import os
import sys
from datetime import datetime
from openpyxl import load_workbook

# Configuration
ITD_FILE = os.environ.get('ITD_FILE', '/Users/jamesprice/mymineralwatch/ITD-wells-formations-base.xlsx')
BATCH_SIZE = 1000
OUTPUT_DIR = '.'

# Parse args
for i, arg in enumerate(sys.argv[1:]):
    if arg == '--itd-file' and i + 2 < len(sys.argv):
        ITD_FILE = sys.argv[i + 2]

print(f"ITD Formation Import Script")
print(f"Time: {datetime.now()}")
print(f"File: {ITD_FILE}")
print()

if not os.path.exists(ITD_FILE):
    print(f"Error: {ITD_FILE} not found!")
    sys.exit(1)

def escape_sql(value):
    """Escape single quotes for SQL"""
    if value is None:
        return None
    return str(value).replace("'", "''")

# Open workbook in read-only mode for streaming
print("Opening ITD file (154MB)...")
wb = load_workbook(filename=ITD_FILE, read_only=True, data_only=True)
ws = wb.active

# Track stats
stats = {
    'total_rows': 0,
    'valid_apis': 0,
    'apis_with_formation': 0,
    'updates_generated': 0,
    'skipped_no_api': 0,
    'skipped_no_formation': 0,
}

# For each API, keep only the FIRST formation encountered (primary target formation)
# ITD file can have multiple rows per API (multiple target formations per permit)
seen_apis = set()
updates_buffer = []
batch_num = 0

def write_batch():
    """Write current batch to SQL file"""
    global batch_num, updates_buffer
    if not updates_buffer:
        return
    batch_num += 1
    filename = f'{OUTPUT_DIR}/itd-formation-batch-{batch_num:04d}.sql'
    with open(filename, 'w') as f:
        f.write('\n'.join(updates_buffer))
    print(f"  Wrote {len(updates_buffer)} statements to {filename}")
    updates_buffer.clear()

# Process rows
for row_num, row in enumerate(ws.iter_rows(values_only=True)):
    if row_num == 0:
        # Verify expected columns
        headers = [str(c) if c else '' for c in row]
        assert 'Formation_Name' in headers[49] or 'Formation' in str(headers[49]), \
            f"Expected Formation_Name at column 49, got: {headers[49]}"
        continue

    stats['total_rows'] += 1
    if stats['total_rows'] % 50000 == 0:
        print(f"  Processed {stats['total_rows']:,} rows...")

    # Get API number - normalize to 10-digit
    api_raw = row[0]
    if not api_raw:
        stats['skipped_no_api'] += 1
        continue

    api_str = ''.join(c for c in str(api_raw) if c.isdigit())
    if len(api_str) < 10:
        stats['skipped_no_api'] += 1
        continue

    api_10 = api_str[:10]

    # Skip if already processed this API
    if api_10 in seen_apis:
        continue
    seen_apis.add(api_10)
    stats['valid_apis'] += 1

    # Get formation name
    form_name = row[49]
    if not form_name or not str(form_name).strip():
        stats['skipped_no_formation'] += 1
        continue

    form_name_clean = str(form_name).strip()
    stats['apis_with_formation'] += 1

    # Build COALESCE UPDATE — only fills NULL formation_name
    escaped_name = escape_sql(form_name_clean)
    sql = (
        f"UPDATE wells SET "
        f"formation_name = COALESCE(formation_name, '{escaped_name}'), "
        f"formation_canonical = COALESCE(formation_canonical, "
        f"(SELECT canonical_name FROM formation_normalization WHERE raw_name = '{escaped_name}')), "
        f"formation_group = COALESCE(formation_group, "
        f"(SELECT formation_group FROM formation_normalization WHERE raw_name = '{escaped_name}')) "
        f"WHERE api_number = '{api_10}' AND formation_name IS NULL;"
    )
    updates_buffer.append(sql)
    stats['updates_generated'] += 1

    if len(updates_buffer) >= BATCH_SIZE:
        write_batch()

# Write final batch
write_batch()
wb.close()

# Print summary
print()
print("=" * 60)
print("ITD Formation Import Complete!")
print(f"Total rows processed: {stats['total_rows']:,}")
print(f"Unique APIs: {stats['valid_apis']:,}")
print(f"APIs with formation name: {stats['apis_with_formation']:,}")
print(f"UPDATE statements generated: {stats['updates_generated']:,}")
print(f"Batch files created: {batch_num}")
print(f"Skipped (no API): {stats['skipped_no_api']:,}")
print(f"Skipped (no formation): {stats['skipped_no_formation']:,}")
print("=" * 60)

# Create execution script
if batch_num > 0:
    script = f"""#!/bin/bash
# Execute ITD formation data import
# Generated: {datetime.now().isoformat()}
# Batches: {batch_num}
# Updates: {stats['updates_generated']:,}
echo "Starting ITD formation import..."
echo "Total batches: {batch_num}"
echo ""

start_time=$(date +%s)
processed=0
errors=0

for file in itd-formation-batch-*.sql; do
    if [ -f "$file" ]; then
        processed=$((processed + 1))
        echo "[$(date '+%H:%M:%S')] Batch $processed/{batch_num}: $file"

        if wrangler d1 execute oklahoma-wells --remote --file="$file"; then
            echo "  OK"
        else
            echo "  FAILED"
            errors=$((errors + 1))
        fi

        # Brief pause between batches
        if [ $processed -lt {batch_num} ]; then
            sleep 2
        fi
    fi
done

end_time=$(date +%s)
duration=$((end_time - start_time))

echo ""
echo "ITD formation import complete!"
echo "Time: $duration seconds"
echo "Batches: $processed"
echo "Errors: $errors"
echo ""
echo "Next steps:"
echo "  1. Verify: SELECT COUNT(*) as total, SUM(CASE WHEN formation_name IS NOT NULL THEN 1 ELSE 0 END) as has_formation FROM wells;"
echo "  2. Check normalization: SELECT COUNT(*) FROM wells WHERE formation_name IS NOT NULL AND formation_group IS NULL;"
echo "  3. Re-run risk profile assignment if needed"
"""
    with open(f'{OUTPUT_DIR}/execute-itd-import.sh', 'w') as f:
        f.write(script)
    os.chmod(f'{OUTPUT_DIR}/execute-itd-import.sh', 0o755)
    print(f"\nCreated: ./execute-itd-import.sh")
    print(f"\nRun migration 022 first, then: ./execute-itd-import.sh")

print(f"\nFinished at: {datetime.now()}")
